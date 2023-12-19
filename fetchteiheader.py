#!/usr/bin/env python3
from datetime import datetime
from acdh_tei_pyutils.tei import TeiReader
from lxml import etree as ET
import json
import os
import re
import pandas as pd
from openpyxl import load_workbook


nsmap = {
    "tei": "http://www.tei-c.org/ns/1.0",
    "mets": "http://www.loc.gov/METS/",
    "mods": "http://www.loc.gov/mods/v3",
    "dv": "http://dfg-viewer.de/",
    "default": "http://www.tei-c.org/ns/1.0",
}


class Log:
    def __init__(self, logfile):
        self.file_name = logfile

    def print_log(self, faulty_item, arg="could not be parsed"):
        now = datetime.now()
        with open(f"{self.file_name}.log", "a") as f:
            f.write(
                f'{now.strftime("%Y/%m/%d %H:%M:%S")}\t' f"File `{faulty_item}` {arg}\n"
            )


log = Log("mets2tei.log")


class TEIValidator:
    from lxml import objectify
    from lxml.etree import XMLSyntaxError, XMLSchema, parse
    from lxml import etree as ET

    def __init__(self, xml_tree, xsd_file):
        self.xsd_root = self.ET.parse(xsd_file)
        self.xml_tree = xml_tree
        self.valid = self.validate()

    def validate(self):
        validated = True
        try:
            schema = self.XMLSchema(self.xsd_root)
            schema.validate(self.xml_tree)
        except self.XMLSyntaxError:
            validated = False
        return validated


class TeiTree:
    def __init__(self, source_table, source_tkb):
        self.wb_obj = load_workbook(source_table).active
        self.doc_id = os.path.basename(source_tkb).rstrip(".xml")
        self.tkb = TeiReader(source_tkb)
        self.tei = TeiReader("template.xml")
        self.root = self.tei.any_xpath("//tei:TEI")[0]
        self.header = self.tei.any_xpath("//tei:teiHeader")[0]
        self.msdesc = self.tei.any_xpath("//tei:msDesc")[0]
        self.elements = self.extract_from_table(source_table, self.header)
        self.root.append(self.tkb.any_xpath("//tei:facsimile")[0])
        self.root.append(self.tkb.any_xpath("//tei:text")[0])

    def extract_from_table(self, table, header):
        self.header = header
        df = pd.read_excel(table).fillna("")
        column_name = self.doc_id[0] + " " + self.doc_id[1:].replace("_", "/")

        for idx, row in df.loc[df["Signatur"] == column_name].iterrows():
            self.parse_signature(row["Signatur"])
            self.parse_origin(row["Provenienz"])
            self.parse_date(str(row["Zeit"]))
            keys = self.classify_books(row["Buchtyp"], row["Liturgie"])
            self.parse_summary(row["Inhalt"], row["Buchtyp"], keys, idx)
            self.parse_extension(row["Umfang fol."])
            self.parse_format(row["Format"])
            self.parse_notation(False)  # Placeholder

    def parse_signature(self, sign):
        # sign = sign.replace("/", "_").replace(" ", "")
        self.msdesc.xpath("//tei:msIdentifier/tei:idno", namespaces=nsmap)[
            0
        ].text = sign
        if not self.header.xpath("//tei:titleStmt/tei:title", namespaces=nsmap)[0].text:
            self.header.xpath("//tei:titleStmt/tei:title", namespaces=nsmap)[
                0
            ].text = sign

    def parse_origin(self, origin):
        self.msdesc.xpath("//tei:history/tei:provenance", namespaces=nsmap)[
            0
        ].text = origin

    def parse_date(self, date):
        element = self.msdesc.xpath(
            "//tei:fileDesc/tei:sourceDesc/tei:bibl/tei:date", namespaces=nsmap
        )[0]
        try:
            year = re.sub("x+", "00", date).lstrip("~").split()[0]
            year = int(year.split("-")[0].strip("."))
        except Exception:
            log.print_log(date, "is not a valid date")
            year = "nan"
        if date.startswith("~"):
            ddate = {"notBefore": f"{year - 20}", "notAfter": f"{year + 20}"}
        elif date.endswith("Jh."):
            ddate = {
                "notBefore": f"{(year - 1) * 100}",
                "notAfter": f"{(year - 1) * 100 + 99}",
            }
        elif date.endswith("x"):
            ddate = {"notBefore": f"{year}", "notAfter": f"{year + 99}"}
        elif re.findall(r"^\d{2}\-\d(?:/\d)*", date):
            second = date.split("-")[1]
            year *= 100
            if second in "12":
                factor = 100 / int(second)
                ddate = {"notBefore": year - factor, "notAfter": year - factor + 50}
            else:
                factor = int(second.split("/")[0]) * 100 / int(second.split("/")[1])
                ddate = {
                    "notBefore": year + factor - 25,
                    "notAfter": year + factor,
                }
        else:
            ddate = {"when": f"{year}"}
        element.text = date
        for time in ddate:
            element.attrib[time] = str(ddate[time])

    @staticmethod
    def classify_books(booktype, lit):
        if lit == lit:
            keys = f"#{lit.lower()}"
        else:
            keys = ""
        books = " ".join(" ".join(booktype.split(",")).split("/")).split()
        with open("booktypes.json", "r") as f:
            dictionary = json.load(f)
        for book in books:
            for booktype in dictionary:
                if booktype in book.lower():
                    keys += f" #{dictionary[booktype]}"
        return keys

    def parse_summary(self, summary, bookt, attributes, line):
        if summary == summary:
            content = self.wb_obj.cell(row=line + 1, column=6)
        else:
            content = ""
        element = self.msdesc.xpath("//tei:msContents", namespaces=nsmap)[0]
        element.attrib["class"] = attributes
        subelement = element.xpath("//tei:summary", namespaces=nsmap)[0]
        ET.SubElement(subelement, "p").text = str(content)
        ET.SubElement(subelement, "p").text = bookt

    def parse_extension(self, umfang):
        tree = self.msdesc.xpath(
            "//tei:physDesc/tei:objectDesc/tei:supportDesc/tei:extent/tei:measure",
            namespaces=nsmap,
        )[0]
        tree.attrib["unit"] = "leaf"
        if isinstance(umfang, (int, float)):
            umfang = str(int(umfang))
            tree.attrib["quantity"] = umfang
        else:
            for i in umfang.split():
                if i.isnumeric():
                    tree.attrib["quantity"] = i
                    break
        tree.text = umfang

    def parse_format(self, size):
        tree = self.msdesc.xpath(
            "//tei:physDesc/tei:objectDesc/tei:supportDesc/tei:support/tei:dimensions",
            namespaces=nsmap,
        )[0]
        size = size.replace("*", "x").split("x")
        if len(size) < 2:
            if size:
                ET.SubElement(tree, "dim").text = size[0]
        else:
            ET.SubElement(tree, "height").text = size[0]
            ET.SubElement(tree, "width").text = size[1]

    def parse_notation(self, placeholder):
        # We don't have the source yet.
        tree = self.msdesc.xpath("//tei:physDesc/tei:musicNotation", namespaces=nsmap)[
            0
        ]
        tree.text = "Placeholder"
        return tree
